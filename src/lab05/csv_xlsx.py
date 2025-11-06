import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    csv_file = Path(csv_path)

    if not csv_file.exists():
        raise FileNotFoundError(f'Файл {csv_file} не найден')
    
    if csv_file.suffix.lower() != '.csv':
        raise ValueError(f'Файл {csv_file} не csv')
    
    xlsx_file = Path(xlsx_path)
    if xlsx_file.suffix.lower() != '.xlsx':
        raise ValueError(f'Файл {xlsx_file} не xlsx')
    
    try:
        with csv_file.open('r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)

    except csv.Error as e:
        raise ValueError(f"Invalid CSV format: {e}")
    except UnicodeDecodeError:
        raise ValueError("CSV file encoding must be UTF-8")
    
    if len(rows) == 0:
        raise ValueError('csv-файл пустой')
    
    if not rows[0] or all(cell.strip() == '' for cell in rows[0]):
        raise ValueError("CSV file must have a header row")
    
    xlsx_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'lst1'

    for row_idx, row in enumerate(rows, 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = max(length + 2, 8)  # +2 для отступов, минимум 8
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(xlsx_path)

csv_to_xlsx('data/out/people.csv', 'data/out/csv_to_xlsx.xlsx')
